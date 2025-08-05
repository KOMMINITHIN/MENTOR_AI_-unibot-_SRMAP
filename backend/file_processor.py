import os
import io
import base64
from typing import Dict, List, Optional, Tuple
from PIL import Image
import PyPDF2
import pytesseract
from docx import Document
import openpyxl
import json
import re

class FileProcessor:
    """Advanced file processing with AI analysis capabilities"""
    
    def __init__(self):
        self.supported_formats = {
            'pdf': self.process_pdf,
            'png': self.process_image,
            'jpg': self.process_image,
            'jpeg': self.process_image,
            'gif': self.process_image,
            'bmp': self.process_image,
            'tiff': self.process_image,
            'docx': self.process_docx,
            'txt': self.process_text,
            'py': self.process_code,
            'js': self.process_code,
            'html': self.process_code,
            'css': self.process_code,
            'java': self.process_code,
            'cpp': self.process_code,
            'c': self.process_code,
            'xlsx': self.process_excel,
            'csv': self.process_csv,
            'json': self.process_json,
            'md': self.process_markdown
        }
    
    def process_file(self, file_content: bytes, filename: str, file_type: str) -> Dict:
        """Main file processing function"""
        try:
            # Get file extension
            ext = filename.split('.')[-1].lower() if '.' in filename else file_type.lower()
            
            if ext not in self.supported_formats:
                return {
                    "success": False,
                    "error": f"Unsupported file format: {ext}",
                    "supported_formats": list(self.supported_formats.keys())
                }
            
            # Process the file
            processor = self.supported_formats[ext]
            result = processor(file_content, filename)
            
            # Add metadata
            result.update({
                "filename": filename,
                "file_type": ext,
                "file_size": len(file_content),
                "success": True
            })
            
            return result
            
        except Exception as e:
            return {
                "success": False,
                "error": f"Error processing file: {str(e)}",
                "filename": filename
            }
    
    def process_pdf(self, content: bytes, filename: str) -> Dict:
        """Extract text from PDF files"""
        try:
            pdf_file = io.BytesIO(content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            
            text_content = ""
            page_count = len(pdf_reader.pages)
            
            for page_num, page in enumerate(pdf_reader.pages):
                text_content += f"\n--- Page {page_num + 1} ---\n"
                text_content += page.extract_text()
            
            return {
                "content_type": "document",
                "text_content": text_content,
                "page_count": page_count,
                "analysis_type": "pdf_extraction",
                "summary": f"PDF document with {page_count} pages extracted successfully"
            }
            
        except Exception as e:
            return {"error": f"PDF processing failed: {str(e)}"}
    
    def process_image(self, content: bytes, filename: str) -> Dict:
        """Process images with OCR for text extraction"""
        try:
            image = Image.open(io.BytesIO(content))
            
            # Convert to RGB if necessary
            if image.mode != 'RGB':
                image = image.convert('RGB')
            
            # Extract text using OCR with enhanced settings
            try:
                # Try multiple OCR configurations for better accuracy
                configs = [
                    '--psm 6',  # Uniform block of text
                    '--psm 4',  # Single column of text
                    '--psm 3',  # Fully automatic page segmentation
                ]

                extracted_text = ""
                for config in configs:
                    try:
                        text = pytesseract.image_to_string(image, config=config)
                        if len(text.strip()) > len(extracted_text.strip()):
                            extracted_text = text
                    except:
                        continue

                if not extracted_text.strip():
                    extracted_text = "No text detected in image"

            except Exception as ocr_error:
                extracted_text = f"OCR processing failed: {str(ocr_error)}"
            
            # Get image info
            width, height = image.size
            
            # Detect if it's likely handwritten (basic heuristic)
            is_handwritten = self.detect_handwriting(extracted_text)
            
            return {
                "content_type": "image",
                "text_content": extracted_text,
                "image_info": {
                    "width": width,
                    "height": height,
                    "format": image.format,
                    "mode": image.mode
                },
                "analysis_type": "ocr_extraction",
                "is_handwritten": is_handwritten,
                "summary": f"Image ({width}x{height}) processed with OCR"
            }
            
        except Exception as e:
            return {"error": f"Image processing failed: {str(e)}"}
    
    def process_docx(self, content: bytes, filename: str) -> Dict:
        """Extract text from Word documents"""
        try:
            doc_file = io.BytesIO(content)
            doc = Document(doc_file)
            
            text_content = ""
            paragraph_count = 0
            
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    text_content += paragraph.text + "\n"
                    paragraph_count += 1
            
            return {
                "content_type": "document",
                "text_content": text_content,
                "paragraph_count": paragraph_count,
                "analysis_type": "docx_extraction",
                "summary": f"Word document with {paragraph_count} paragraphs extracted"
            }
            
        except Exception as e:
            return {"error": f"DOCX processing failed: {str(e)}"}
    
    def process_text(self, content: bytes, filename: str) -> Dict:
        """Process plain text files"""
        try:
            text_content = content.decode('utf-8')
            line_count = len(text_content.split('\n'))
            word_count = len(text_content.split())
            
            return {
                "content_type": "text",
                "text_content": text_content,
                "line_count": line_count,
                "word_count": word_count,
                "analysis_type": "text_extraction",
                "summary": f"Text file with {line_count} lines and {word_count} words"
            }
            
        except Exception as e:
            return {"error": f"Text processing failed: {str(e)}"}
    
    def process_code(self, content: bytes, filename: str) -> Dict:
        """Process code files with syntax analysis"""
        try:
            code_content = content.decode('utf-8')
            ext = filename.split('.')[-1].lower()
            
            # Basic code analysis
            line_count = len(code_content.split('\n'))
            
            # Detect programming language features
            language_features = self.analyze_code_language(code_content, ext)
            
            return {
                "content_type": "code",
                "text_content": code_content,
                "programming_language": ext,
                "line_count": line_count,
                "language_features": language_features,
                "analysis_type": "code_analysis",
                "summary": f"{ext.upper()} code file with {line_count} lines analyzed"
            }
            
        except Exception as e:
            return {"error": f"Code processing failed: {str(e)}"}
    
    def process_excel(self, content: bytes, filename: str) -> Dict:
        """Process Excel files"""
        try:
            excel_file = io.BytesIO(content)
            workbook = openpyxl.load_workbook(excel_file)
            
            sheets_data = {}
            total_rows = 0
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        sheet_data.append([str(cell) if cell is not None else "" for cell in row])
                        total_rows += 1
                
                sheets_data[sheet_name] = sheet_data
            
            # Convert to text representation
            text_content = self.excel_to_text(sheets_data)
            
            return {
                "content_type": "spreadsheet",
                "text_content": text_content,
                "sheets_data": sheets_data,
                "sheet_count": len(workbook.sheetnames),
                "total_rows": total_rows,
                "analysis_type": "excel_extraction",
                "summary": f"Excel file with {len(workbook.sheetnames)} sheets and {total_rows} rows"
            }
            
        except Exception as e:
            return {"error": f"Excel processing failed: {str(e)}"}
    
    def process_csv(self, content: bytes, filename: str) -> Dict:
        """Process CSV files"""
        try:
            text_content = content.decode('utf-8')
            lines = text_content.split('\n')
            row_count = len([line for line in lines if line.strip()])
            
            return {
                "content_type": "data",
                "text_content": text_content,
                "row_count": row_count,
                "analysis_type": "csv_extraction",
                "summary": f"CSV file with {row_count} rows processed"
            }
            
        except Exception as e:
            return {"error": f"CSV processing failed: {str(e)}"}
    
    def process_json(self, content: bytes, filename: str) -> Dict:
        """Process JSON files"""
        try:
            text_content = content.decode('utf-8')
            json_data = json.loads(text_content)
            
            # Pretty format JSON
            formatted_json = json.dumps(json_data, indent=2)
            
            return {
                "content_type": "data",
                "text_content": formatted_json,
                "json_structure": self.analyze_json_structure(json_data),
                "analysis_type": "json_extraction",
                "summary": f"JSON file with structured data analyzed"
            }
            
        except Exception as e:
            return {"error": f"JSON processing failed: {str(e)}"}
    
    def process_markdown(self, content: bytes, filename: str) -> Dict:
        """Process Markdown files"""
        try:
            text_content = content.decode('utf-8')
            
            # Basic markdown analysis
            headers = re.findall(r'^#+\s+(.+)$', text_content, re.MULTILINE)
            links = re.findall(r'\[([^\]]+)\]\(([^)]+)\)', text_content)
            
            return {
                "content_type": "document",
                "text_content": text_content,
                "headers": headers,
                "links": links,
                "analysis_type": "markdown_extraction",
                "summary": f"Markdown file with {len(headers)} headers analyzed"
            }
            
        except Exception as e:
            return {"error": f"Markdown processing failed: {str(e)}"}
    
    def detect_handwriting(self, text: str) -> bool:
        """Basic heuristic to detect if text might be handwritten"""
        if not text or len(text.strip()) < 10:
            return False
        
        # Look for OCR artifacts that suggest handwriting
        handwriting_indicators = [
            len(re.findall(r'[^\w\s]', text)) / len(text) > 0.1,  # High special char ratio
            len(text.split()) / len(text.split('\n')) < 3,  # Few words per line
            any(word.islower() and len(word) == 1 for word in text.split())  # Single lowercase letters
        ]
        
        return sum(handwriting_indicators) >= 2
    
    def analyze_code_language(self, code: str, ext: str) -> Dict:
        """Analyze code for language-specific features"""
        features = {
            "functions": len(re.findall(r'def\s+\w+|function\s+\w+|void\s+\w+|int\s+\w+', code)),
            "classes": len(re.findall(r'class\s+\w+', code)),
            "imports": len(re.findall(r'import\s+|#include\s+|require\s+', code)),
            "comments": len(re.findall(r'//.*|#.*|/\*.*?\*/', code, re.DOTALL))
        }
        return features
    
    def excel_to_text(self, sheets_data: Dict) -> str:
        """Convert Excel data to readable text"""
        text_parts = []
        for sheet_name, data in sheets_data.items():
            text_parts.append(f"=== Sheet: {sheet_name} ===")
            for row in data[:10]:  # Limit to first 10 rows
                text_parts.append(" | ".join(row))
            if len(data) > 10:
                text_parts.append(f"... and {len(data) - 10} more rows")
        return "\n".join(text_parts)
    
    def analyze_json_structure(self, data) -> Dict:
        """Analyze JSON structure"""
        if isinstance(data, dict):
            return {"type": "object", "keys": list(data.keys())[:10]}
        elif isinstance(data, list):
            return {"type": "array", "length": len(data)}
        else:
            return {"type": type(data).__name__}
